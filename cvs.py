import sys
from openpyxl import load_workbook
import csv

wb = load_workbook("./qbInvoiceImport.xlsx")
solsht = wb.worksheets[0]
propsht = wb.worksheets[1]
prodsht = wb.worksheets[2]
inputsht = wb.worksheets[3]

proprows = tuple(propsht.iter_rows(min_row=2,min_col=1,max_col=2,values_only=True))
properties = dict(proprows)
prodrows = tuple(prodsht.iter_rows(min_row=2,min_col=1,max_col=2,values_only=True))
products = dict(prodrows)

# get maximum row in input sheet
numrows = inputsht.max_row
# variables for building, check number and check date
bldg = ""
tmpbldg = ""
invoicenum = int(sys.argv[1])
print(invoicenum)
ckdate = inputsht.cell(row=1,column=2).value
ckdate = ckdate.rstrip()
# list used for outputing to csv
row_list = []

# go through each row in input sheet and get property, product and amount
for i in range(3,numrows):
    previousrw = i - 1
    col_list=[]
    bldg = inputsht.cell(row=i, column=2).value
    glacc = inputsht.cell(row=i, column=4).value
    amnt = inputsht.cell(row=i,column=5).value

    if i <= numrows:
        tmpbldg = inputsht.cell(row=previousrw, column=2).value
    if tmpbldg == bldg:
        prop = ""
    else:
        # find property in dictionary
        prop = properties.get(bldg)
        invoicenum = invoicenum + 1

    # find product in dictionary
    product = products.get(glacc)

    # put fields into list
    col_list.append(invoicenum)

    if prop == "":
        col_list.append('')
        col_list.append('')
        col_list.append('')
    else:
        col_list.append(prop)
        col_list.append(ckdate)
        col_list.append(ckdate)
    col_list.append('')
    col_list.append('')
    col_list.append('')
    col_list.append(product)
    col_list.append('')
    col_list.append(1)
    col_list.append('')
    col_list.append(amnt)
    row_list.append(col_list)

with open('test.csv', 'w', newline='') as file:
    writer = csv.writer(file)
    writer.writerows(row_list)







